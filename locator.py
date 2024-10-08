import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime

# Excel file and sheet setup
excel_file = "test_data.xlsx"  # Path to your Excel file
sheet_number = 0  # Assuming it's the first sheet

# Web Data Constants
class WebData:
    url = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"
    dashboard_url = "https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index"

# Excel Helper Functions
class ExcelFunctions:
    def __init__(self, file, sheet_number):
        self.file = file
        self.wb = openpyxl.load_workbook(file)
        self.sheet = self.wb.worksheets[sheet_number]

    def row_count(self):
        return self.sheet.max_row

    def read_data(self, row, col):
        return self.sheet.cell(row=row, column=col).value

    def write_data(self, row, col, value):
        self.sheet.cell(row=row, column=col).value = value
        self.wb.save(self.file)

    def update_test_time(self, row, date_col, time_col):
        current_time = datetime.now()
        date_str = current_time.strftime("%Y-%m-%d")
        time_str = current_time.strftime("%I:%M %p")
        self.write_data(row, date_col, date_str)
        self.write_data(row, time_col, time_str)

# Page Object Model (POM) - Login Locators
class LoginPageLocators:
    USERNAME = (By.NAME, "username")
    PASSWORD = (By.NAME, "password")
    SUBMIT_BUTTON = (By.XPATH, "//button[@type='submit']")
    DASHBOARD = (By.XPATH, "//h6[text()='Dashboard']")
    LOGOUT_BUTTON = (By.XPATH, "//a[@href='/web/index.php/auth/logout']")

# Setup WebDriver
def setup_driver():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.maximize_window()
    driver.get(WebData.url)
    return driver

# Test Login
def test_login():
    # Load Excel File
    excel = ExcelFunctions(excel_file, sheet_number)
    driver = setup_driver()

    rows = excel.row_count()

    for row in range(2, rows + 1):
        username = excel.read_data(row, 2)  # Username in column 2
        password = excel.read_data(row, 3)  # Password in column 3

        # Enter username and password
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(LoginPageLocators.USERNAME)
        ).send_keys(username)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(LoginPageLocators.PASSWORD)
        ).send_keys(password)

        # Click login button
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(LoginPageLocators.SUBMIT_BUTTON)
        ).click()

        # Check for successful login by verifying dashboard presence
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(LoginPageLocators.DASHBOARD)
            )
            print(f"SUCCESS: Login successful for {username}")
            excel.write_data(row, 7, "TEST PASS")  # Write 'PASS' in column 7
        except:
            print(f"FAIL: Login failed for {username}")
            excel.write_data(row, 7, "TEST FAIL")  # Write 'FAIL' in column 7

        # Update date and time in Excel file (Date in column 4, Time in column 5)
        excel.update_test_time(row, 4, 5)

        # Go back to login page for the next test
        driver.get(WebData.url)

    driver.quit()

# Run the test
if __name__ == "__main__":
    test_login()
